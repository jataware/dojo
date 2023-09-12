from __future__ import annotations

import codecs
import csv
import io
import json
import os
import re
import time
import uuid
from datetime import datetime
from functools import partial
from typing import List, Optional
from urllib.parse import urlparse

import openpyxl
import pandas as pd
from elasticsearch import Elasticsearch
from fastapi import APIRouter, File, HTTPException, Query, Response, UploadFile, status, Request
from fastapi.logger import logger
from fastapi.responses import FileResponse
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError
from redis import Redis
from rq import Queue
from src.causemos import deprecate_dataset, notify_causemos
from src.csv_annotation_parser import format_annotations, xls_to_annotations
from src.data import get_context, job
from src.dojo import search_and_scroll
from src.feature_queries import keyword_query_v1, hybrid_query_v2
from src.ontologies import get_ontologies
from src.plugins import plugin_action
from src.settings import settings
from src.utils import (
    add_date_to_dataset, get_rawfile, list_files,
    put_rawfile, format_hybrid_results
)
from validation import DojoSchema, IndicatorSchema, MetadataSchema

router = APIRouter()
es = Elasticsearch([settings.ELASTICSEARCH_URL], port=settings.ELASTICSEARCH_PORT)

# REDIS CONNECTION AND QUEUE OBJECTS
redis = Redis(
    os.environ.get("REDIS_HOST", "redis.dojo-stack"),
    int(os.environ.get("REDIS_PORT", 6379)),
)
q = Queue(connection=redis, default_timeout=-1)


# For created_at times in epoch milliseconds
def current_milli_time():
    return round(time.time() * 1000)


def extract_protocol_host_port(url):
    parsed_url = urlparse(url)
    protocol = parsed_url.scheme
    host = parsed_url.hostname
    port = parsed_url.port

    # Include the port in the result if it's present in the URL
    if port:
        return f"{protocol}://{host}:{port}"
    else:
        return f"{protocol}://{host}"


def enqueue_indicator_feature(indicator_id, indicator_dict):
    """
    Adds indicator (dataset!) to queue to process by embeddings_process, which
    at this time creates and attaches LLM embeddings to its features (outputs)
    """
    job_string = "embeddings_processors.calculate_store_embeddings"
    job_id = f"{indicator_id}_{job_string}"

    context = {"indicator_id": indicator_id, "full_indicator": indicator_dict}

    job = q.enqueue_call(func=job_string, args=[context], kwargs={}, job_id=job_id)


@router.post("/indicators")
def create_indicator(payload: IndicatorSchema.IndicatorMetadataSchema):
    indicator_id = str(uuid.uuid4())
    payload.id = indicator_id
    payload.created_at = current_milli_time()
    body = payload.json()
    payload.published = False

    plugin_action("before_create", data=body, type="indicator")
    es.index(index="indicators", body=body, id=indicator_id)
    plugin_action("post_create", data=body, type="indicator")

    empty_annotations_payload = MetadataSchema.MetaModel(metadata={}).json()
    # (?): SHOULD WE HAVE PLUGINS AROUND THE ANNOTATION CREATION?
    plugin_action("before_create", data=body, type="annotation")
    es.index(index="annotations", body=empty_annotations_payload, id=indicator_id)
    plugin_action("post_create", data=body, type="annotation")

    # Saves all outputs, as features in a new index, with LLM embeddings
    if payload.outputs:
        enqueue_indicator_feature(indicator_id, json.loads(payload.json()))

    return Response(
        status_code=status.HTTP_201_CREATED,
        headers={
            "location": f"/api/indicators/{indicator_id}",
            "content-type": "application/json",
        },
        content=body,
    )


@router.put("/indicators")
def update_indicator(payload: IndicatorSchema.IndicatorMetadataSchema):
    indicator_id = payload.id
    payload.created_at = current_milli_time()
    body = payload.json()

    plugin_action("before_update", data=body, type="indicator")
    es.index(index="indicators", body=body, id=indicator_id)
    plugin_action("post_update", data=body, type="indicator")

    if payload.outputs:
        enqueue_indicator_feature(indicator_id, json.loads(payload.json()))

    return Response(
        status_code=status.HTTP_200_OK,
        headers={"location": f"/api/indicators/{indicator_id}"},
        content=f"Updated indicator with id = {indicator_id}",
    )


@router.patch("/indicators")
def patch_indicator(
    payload: IndicatorSchema.IndicatorMetadataSchema, indicator_id: str
):
    payload.created_at = current_milli_time()
    body = json.loads(payload.json(exclude_unset=True))
    es.update(index="indicators", body={"doc": body}, id=indicator_id)

    updated = es.get_source(
        index="indicators", id=indicator_id, params={"_source": "name,outputs"}
    )
    if updated["outputs"]:
        enqueue_indicator_feature(indicator_id, updated)

    return Response(
        status_code=status.HTTP_200_OK,
        headers={"location": f"/api/indicators/{indicator_id}"},
        content=f"Updated indicator with id = {indicator_id}",
    )


def format_one_result(r):
        r["_source"]["metadata"] = {}
        r["_source"]["metadata"]["match_score"] = r["_score"]
        r["_source"]["id"] = r["_id"]
        r["_source"]["metadata"]["matched_queries"] = r["matched_queries"]
        return r["_source"]


@router.get(
    "/features/search", response_model=IndicatorSchema.FeaturesSemanticSearchSchema
)
def semantic_search_features(
    query: Optional[str], size=10, scroll_id: Optional[str] = None
):
    """
    Given a text query, uses semantic search engine to search for features that
    match the query semantically. Query is a sentence that can be interpreted
    to be related to a concept, such as:
    'number of people who have been vaccinated'
    """

    if scroll_id:
        results = es.scroll(scroll_id=scroll_id, scroll="2m")
    else:
        # Retrieve first item in output, since it returns an array output
        # that matches its input, and we provide only one- query.
        features_query = hybrid_query_v2(query)

        results = es.search(
            index="features", body=features_query, scroll="2m", size=size
        )

    items_in_page = len(results["hits"]["hits"])

    if items_in_page < int(size):
        scroll_id = None
    else:
        scroll_id = results.get("_scroll_id", None)

    max_score = results["hits"]["max_score"]

    first = results["hits"]["hits"][0]

    alternated = format_hybrid_results(results["hits"]["hits"])

    return {
        "hits": results["hits"]["total"]["value"],
        "items_in_page": items_in_page,
        "max_score": max_score,
        "results": [format_one_result(i) for i in alternated],
        "scroll_id": scroll_id,
    }


@router.get("/features", response_model=IndicatorSchema.FeaturesSearchSchema)
def list_features(
    term: Optional[str] = None, size: int = 10, scroll_id: Optional[str] = None
):
    """
    Lists all features, with pagination, or results from searching
    through them (using input sentence `term`). Will match `term`
    to feature `name`, `display_name`, or `description`.
    """

    if term:
        q = keyword_query_v1(term)

    else:
        q = {"query": {"match_all": {}}, "_source": {"excludes": "embeddings"}}

    if not scroll_id:
        results = es.search(index="features", body=q, scroll="2m", size=size)
    else:
        results = es.scroll(scroll_id=scroll_id, scroll="2m")

    es_hits = results["hits"]["hits"]
    hits_count = len(es_hits)

    # if hits are less than the page size (10) don't return a scroll_id
    if hits_count < size:
        scroll_id = None
    else:
        scroll_id = results.get("_scroll_id", None)

    max_score = results["hits"]["max_score"]

    def formatOneResult(r):
        if term:
            r["_source"]["metadata"] = {}
            r["_source"]["metadata"]["match_score"] = r["_score"]
        r["_source"]["id"] = r["_id"]
        return r["_source"]

    response = {
        "hits": results["hits"]["total"]["value"],
        "items_in_page": hits_count,
        "results": [formatOneResult(i) for i in es_hits],
        "scroll_id": scroll_id,
    }

    if term:
        response["max_score"] = max_score

    return response


@router.get(
    "/indicators/latest", response_model=List[IndicatorSchema.IndicatorsSearchSchema]
)
def get_latest_indicators(size=10000):
    q = {
        "_source": [
            "description",
            "name",
            "id",
            "created_at",
            "deprecated",
            "maintainer.name",
            "maintainer.email",
        ],
        "query": {
            "bool": {
                "must": [{"match_all": {}}],
                "filter": [{"term": {"published": True}}],
            }
        },
    }
    results = es.search(index="indicators", body=q, size=size)["hits"]["hits"]
    IndicatorsSchemaArray = []
    for res in results:
        IndicatorsSchemaArray.append(res.get("_source"))
    return IndicatorsSchemaArray

@router.get("/indicators/ncfiles", response_model=List[IndicatorSchema.IndicatorsSearchSchema])
def get_nc_file_indicators(size=10000):
    # match just .nc files
    query = {
        "_source": [
            "description",
            "name",
            "id",
            "created_at",
            "deprecated",
            "maintainer.name",
            "maintainer.email",
            "fileData.raw.url",
            "fileData.raw.rawFileName",
        ],
        "query": {
            "bool": {
                "must": [
                    {"match_all": {}},
                    {"regexp": {
                        "fileData.raw.url": ".*\\.nc"
                    }},
                    {"regexp": {
                        "fileData.raw.rawFileName": ".*\\.nc"
                    }},
                ],
                "filter": [
                    {"term": {"published": True}}
                ]
            }
        }
    }

    results = es.search(index="indicators", body=query, size=size)["hits"]["hits"]
    IndicatorsSchemaArray = []
    for res in results:
        IndicatorsSchemaArray.append(res.get("_source"))

    return IndicatorsSchemaArray


@router.get("/indicators", response_model=DojoSchema.IndicatorSearchResult)
def search_indicators(
    query: str = Query(None),
    size: int = 10,
    scroll_id: str = Query(None),
    include_ontologies: bool = True,
    include_geo: bool = True,
) -> DojoSchema.IndicatorSearchResult:
    indicator_data = search_and_scroll(
        index="indicators", size=size, query=query, scroll_id=scroll_id
    )
    # if request wants ontologies and geo data return all
    if include_ontologies and include_geo:
        return indicator_data
    else:
        for indicator in indicator_data["results"]:
            if not include_ontologies:
                for q_output in indicator["qualifier_outputs"]:
                    try:
                        q_output["ontologies"] = {
                            "concepts": None,
                            "processes": None,
                            "properties": None,
                        }
                    except Exception as e:
                        print(e)
                        logger.exception(e)
                for outputs in indicator["outputs"]:
                    try:
                        outputs["ontologies"] = {
                            "concepts": None,
                            "processes": None,
                            "properties": None,
                        }
                    except Exception as e:
                        print(e)
                        logger.exception(e)
            if not include_geo:
                indicator["geography"]["country"] = []
                indicator["geography"]["admin1"] = []
                indicator["geography"]["admin2"] = []
                indicator["geography"]["admin3"] = []

        return indicator_data


# Comes before the /indicators/{indicator_id} endpoint so that the `register`
# path here isn't captured by {indicator_id}
@router.get("/indicators/register")
def full_dataset_register_help(request: Request):
    """
    Help for dataset registration endpoint.
    """
    api_url = extract_protocol_host_port(str(request.url))

    return {
        "message": "Use the following URIs to download metadata and dictionary templates.",
        "metadata_template_url": f"{api_url}/indicators/register/template",
        "dictionary_template_url": f"{api_url}/indicators/annotations/file-template",
        "docs": "https://www.dojo-modeling.com/data-registration.html"
    }


@router.get(
    "/indicators/{indicator_id}", response_model=IndicatorSchema.IndicatorMetadataSchema
)
def get_indicators(indicator_id: str) -> IndicatorSchema.IndicatorMetadataSchema:
    """Get Indicator

    Args:
        indicator_id (str): The UUID of the dataset to retrieve from elasticsearch.

    Raises:
        HTTPException: This is raised if the dataset is not found in elasticsearch.

    Returns:
        IndicatorSchema.IndicatorMetadataSchema: Returns the ydantic schema for the dataset that contains a metadata dictionary.
    """
    try:
        indicator = es.get(index="indicators", id=indicator_id)["_source"]
    except Exception as e:
        logger.exception(e)
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    return indicator


@router.put("/indicators/{indicator_id}/publish")
def publish_indicator(indicator_id: str):
    try:
        # Update indicator model with ontologies from UAZ
        indicator = es.get(index="indicators", id=indicator_id)["_source"]
        indicator["published"] = True
        # data = get_ontologies(indicator, type="indicator")
        # es.index(index="indicators", body=data, id=indicator_id)

        # Notify Causemos that an indicator was created
        plugin_action("before_publish", data=indicator, type="indicator")
        # TODO: Move notify_causemos only to causemos plugin
        # notify_causemos(data, type="indicator")
        plugin_action("publish", data=indicator, type="indicator")
        plugin_action("post_publish", data=indicator, type="indicator")
    except Exception as e:
        logger.exception(e)
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    return Response(
        status_code=status.HTTP_200_OK,
        headers={"location": f"/api/indicators/{indicator_id}/publish"},
        content=f"Published indicator with id {indicator_id}",
    )


@router.put("/indicators/{indicator_id}/deprecate")
def deprecate_indicator(indicator_id: str):
    try:
        indicator = es.get(index="indicators", id=indicator_id)["_source"]
        indicator["deprecated"] = True
        es.index(index="indicators", id=indicator_id, body=indicator)

        # Tell Causemos to deprecate the dataset on their end
        deprecate_dataset(indicator_id)
    except Exception as e:
        logger.exception(e)
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    return Response(
        status_code=status.HTTP_200_OK,
        headers={"location": f"/api/indicators/{indicator_id}"},
        content=f"Deprecated indicator with id {indicator_id}",
    )


@router.get(
    "/indicators/{indicator_id}/annotations", response_model=MetadataSchema.MetaModel
)
def get_annotations(indicator_id: str) -> MetadataSchema.MetaModel:
    """Get annotations for a dataset.

    Args:
        indicator_id (str): The UUID of the dataset to retrieve annotations for from elasticsearch.

    Raises:
        HTTPException: This is raised if no annotation is found for the dataset in elasticsearch.

    Returns:
        MetadataSchema.MetaModel: Returns the annotations pydantic schema for the dataset that contains a metadata dictionary and an annotations object validated via a nested pydantic schema.
    """
    try:
        annotation = es.get(index="annotations", id=indicator_id)["_source"]
        return annotation
    except Exception as e:
        logger.exception(e)
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)


@router.post("/indicators/{indicator_id}/annotations")
def post_annotation(payload: MetadataSchema.MetaModel, indicator_id: str):
    """Post annotations for a dataset.

    Args:
        payload (MetadataSchema.MetaModel): Payload needs to be a fully formed json object representing the pydantic schema MettaDataSchema.MetaModel.
        indicator_id (str): The UUID of the dataset to retrieve annotations for from elasticsearch.

    Returns:
        Response: Returns a response with the status code of 201 and the location of the annotation.
    """
    try:
        body = json.loads(payload.json())

        es.index(index="annotations", body=body, id=indicator_id)

        return Response(
            status_code=status.HTTP_201_CREATED,
            headers={"location": f"/api/annotations/{indicator_id}"},
            content=f"Updated annotation with id = {indicator_id}",
        )
    except Exception as e:
        logger.exception(e)
        return Response(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content=f"Could not update annotation with id = {indicator_id}",
        )


@router.put("/indicators/{indicator_id}/annotations")
def put_annotation(payload: MetadataSchema.MetaModel, indicator_id: str):
    """Put annotation for a dataset to Elasticsearch.

    Args:
        payload (MetadataSchema.MetaModel): Payload needs to be a fully formed json object representing the pydantic schema MettaDataSchema.MetaModel.
        indicator_id (str): The UUID of the dataset for which the annotations apply.

    Returns:
        Response: Response object with status code, informational messages, and content.
    """
    try:
        body = json.loads(payload.json())

        es.index(index="annotations", body=body, id=indicator_id)

        return Response(
            status_code=status.HTTP_201_CREATED,
            headers={"location": f"/api/annotations/{indicator_id}"},
            content=f"Created annotation with id = {indicator_id}",
        )
    except Exception as e:
        logger.exception(e)

        return Response(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content=f"Could not create annotation with id = {indicator_id}",
        )


@router.patch("/indicators/{indicator_id}/annotations")
def patch_annotation(payload: MetadataSchema.MetaModel, indicator_id: str):
    """Patch annotation for a dataset to Elasticsearch.

    Args:
        payload (MetadataSchema.MetaModel): Payload needs to be a partially formed json object valid for the pydantic schema MettaDataSchema.MetaModel.
        indicator_id (str): The UUID of the dataset for which the annotations apply.

    Returns:
        Response: Response object with status code, informational messages, and content.
    """
    try:
        body = json.loads(payload.json(exclude_unset=True))

        # Handles datasets being regsitered with no date.
        if body.get("annotations"):
            if not body.get("annotations").get("date", []):
                logger.info("No Date Annotated, making one.")
                rawfile_path = os.path.join(
                    settings.DATASET_STORAGE_BASE_URL,
                    indicator_id,
                    f"raw_data.csv",
                )
                date_annotation = add_date_to_dataset(path=rawfile_path)
                body["annotations"]["date"] = [date_annotation]

            # Check if there is a primary date
            primary_set = False
            for date in body.get("annotations").get("date"):
                if date.get("primary_date"):
                    primary_set = True
                    break
            # If there isn't set the first date to primary
            if not primary_set:
                body["annotations"]["date"][0]["primary_date"] = True

        es.update(index="annotations", body={"doc": body}, id=indicator_id)

        return Response(
            status_code=status.HTTP_200_OK,
            headers={"location": f"/api/annotations/{indicator_id}"},
            content=f"Updated annotation with id = {indicator_id}",
        )
    except Exception as e:
        logger.exception(e)
        return Response(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content=f"Could not update annotation with id = {indicator_id}",
        )


@router.post("/indicators/{indicator_id}/upload")
def upload_file(
    indicator_id: str,
    file: UploadFile = File(...),
    filename: Optional[str] = None,
    append: Optional[bool] = False,
):
    original_filename = file.filename
    _, ext = os.path.splitext(original_filename)
    dir_path = os.path.join(settings.DATASET_STORAGE_BASE_URL, indicator_id)
    if filename is None:
        if append:
            filenum = len(
                [
                    f
                    for f in list_files(dir_path)
                    if os.path.basename(f).startswith("raw_data") and f.endswith(ext)
                ]
            )
            filename = f"raw_data_{filenum}{ext}"
        else:
            filename = f"raw_data{ext}"

    # Upload file
    dest_path = os.path.join(settings.DATASET_STORAGE_BASE_URL, indicator_id, filename)
    put_rawfile(path=dest_path, fileobj=file.file)

    return Response(
        status_code=status.HTTP_201_CREATED,
        headers={
            "location": f"/api/indicators/{indicator_id}",
            "content-type": "application/json",
        },
        content=json.dumps({"id": indicator_id, "filename": filename}),
    )


@router.get("/indicators/{indicator_id}/verbose")
def get_all_indicator_info(indicator_id: str):
    indicator = get_indicators(indicator_id)
    annotations = get_annotations(indicator_id)

    verbose_return_object = {"indicators": indicator, "annotations": annotations}

    return verbose_return_object


@router.post(
    "/indicators/validate_date",
    response_model=IndicatorSchema.DateValidationResponseSchema,
)
def validate_date(payload: IndicatorSchema.DateValidationRequestSchema):
    valid = True
    try:
        for value in payload.values:
            datetime.strptime(value, payload.format)
    except ValueError as e:
        logger.exception(e)
        valid = False

    return {
        "format": payload.format,
        "valid": valid,
    }


@router.post("/indicators/{indicator_id}/preview/{preview_type}")
async def create_preview(
    indicator_id: str,
    preview_type: IndicatorSchema.PreviewType,
    filename: Optional[str] = Query(None),
    filepath: Optional[str] = Query(None),
):
    """Get preview for a dataset.

    Args:
        indicator_id (str): The UUID of the dataset to return a preview of.

    Returns:
        JSON: Returns a json object containing the preview for the dataset.
    """
    try:
        if filename:
            file_suffix_match = re.search(r"raw_data(_\d+)?\.", filename)
            if file_suffix_match:
                file_suffix = file_suffix_match.group(1) or ""
            else:
                file_suffix = ""
        else:
            file_suffix = ""
        # TODO - Get all potential string files concatenated together using list file utility
        if preview_type == IndicatorSchema.PreviewType.processed:
            if filepath:
                rawfile_path = os.path.join(
                    settings.DATASET_STORAGE_BASE_URL,
                    filepath.replace(".csv", ".parquet.gzip"),
                )
            else:
                rawfile_path = os.path.join(
                    settings.DATASET_STORAGE_BASE_URL,
                    indicator_id,
                    f"{indicator_id}{file_suffix}.parquet.gzip",
                )
            file = get_rawfile(rawfile_path)
            df = pd.read_parquet(file)
            try:
                strparquet_path = os.path.join(
                    settings.DATASET_STORAGE_BASE_URL,
                    indicator_id,
                    f"{indicator_id}_str{file_suffix}.parquet.gzip",
                )
                file = get_rawfile(strparquet_path)
                df_str = pd.read_parquet(file)
                df = pd.concat([df, df_str])
            except FileNotFoundError:
                pass

        else:
            if filepath:
                rawfile_path = os.path.join(settings.DATASET_STORAGE_BASE_URL, filepath)
            else:
                rawfile_path = os.path.join(
                    settings.DATASET_STORAGE_BASE_URL, indicator_id, "raw_data.csv"
                )
            file = get_rawfile(rawfile_path)
            df = pd.read_csv(file, delimiter=",")

        obj = json.loads(
            df.sort_index().reset_index(drop=True).head(100).to_json(orient="index")
        )
        indexed_rows = [{"__id": key, **value} for key, value in obj.items()]

        return indexed_rows
    except FileNotFoundError as e:
        logger.exception(e)
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        logger.exception(e)
        return Response(
            status_code=status.HTTP_400_BAD_REQUEST,
            headers={"msg": f"Error: {e}"},
            content=f"Queue could not be deleted.",
        )


@router.put("/indicators/{indicator_id}/rescale")
def rescale_indicator(indicator_id: str):
    from src.data import job

    job_string = "elwood_processors.scale_features"

    resp = job(uuid=indicator_id, job_string=job_string)

    return resp


def get_file_extension(filename):
    # Match the extension part of the filename using a regex pattern
    match = re.search(r'\.([^.]+)$', filename)

    # If there's a match, return the extension; otherwise, return None
    if match:
        return match.group(1)
    else:
        return None


def augment_errors(errors, indicator_id=None, host="http://localhost:8000"):
    """
    Helper Error Fn for full_dataset_register
    """
    errors["doc"] = {
        "dictionary_template_url": f"{host}/indicators/annotations/file-template"
    }
    if indicator_id:
        errors["dataset_id"] = indicator_id
        errors["actions"] = "Please fix all problems and reissue the request by appending the dataset_id from this error detail as an id query parameter. New URL Example: /indicators/register?id=<uuid>."
        errors["doc"] = {
            "dictionary_template_url": f"{host}/indicators/annotations/file-template?id={indicator_id}"
        }

    raise HTTPException(status_code=400, detail=errors)


@router.post("/indicators/register")
async def full_dataset_register(
        request: Request,
        data: UploadFile = File(...),
        metadata: UploadFile = File(...),
        dictionary: UploadFile = File(...),
        id: Optional[str] = Query(None, description="Dataset ID to continue previous failed register attempts."),
):
    """
    Endpoint to register a dataset using one API call. This is different than
    the rest of the endpoints, which requires multiple steps for an API caller
    in order to register a normalized dataset.

    Inputs:
    - _Metadata_ is a json file which contains dataset name, maintainer, etc.
    - _data_ is the actual original dataset file (csv, xls, netcdf, geotiff)
    - _dictionary_ is a xls or csv file that describes/annotates the data's content

    Example create curl:
    curl -v -F "metadata=@metadata.json" -F "data=@data.csv" -F "dictionary=@dictionary.csv" http://dojo-api-host/indicators/register

    If you run into errors and receive a dataset_id on the error details, you may use the id to continue the process after fixing erros:
    curl -v -F "metadata=@metadata.json" -F "data=@data.csv" -F "dictionary=@dictionary.csv" http://dojo-api-host/indicators/register?id=<uuid-from-error>
    This will help ensure there are no dangling datasets and registration is finished.
    """

    completed = []
    api_url = extract_protocol_host_port(str(request.url))

    try:
        # Step 1: Create or Update indicator
        metadata_contents = json.load(metadata.file)
        indicator_metadata = {k: metadata_contents[k] for k in metadata_contents.keys() - {'file_metadata'}}
        is_updating = bool(id)

        if is_updating:
            indicator = IndicatorSchema.IndicatorMetadataSchema.parse_obj({
                **indicator_metadata,
                "id": id
            })
            update_indicator(indicator)
            indicator_body = get_indicators(indicator_id=id)
            indicator_id = id
            completed.append("updated")
        else:
            indicator = IndicatorSchema.IndicatorMetadataSchema.parse_obj(indicator_metadata)
            create_response = create_indicator(indicator)
            indicator_body = json.loads(create_response.body)
            indicator_id = indicator_body["id"]
            completed.append("created")

        # Step 2: Upload dictionary file.
        await upload_data_dictionary_file(indicator_id=indicator_id, file=dictionary)

        completed.append("dictionary")

        # Step 3: Upload raw data file to S3, before converting to xls or anything else
        # (so that rqworker can download and continue the process)
        upload_file(indicator_id=indicator_id, file=data)

        completed.append("s3")

        ext_mapping = {
            "xls": "excel",
            "xlsx": "excel",
            "nc": "netcdf",
            "csv": "csv",
            "tif": "geotiff",
        }

        extension = get_file_extension(data.filename)
        raw_filename = f"raw_data.{extension}"

        annotation_payload = MetadataSchema.MetaModel(metadata={
            "files": {
                raw_filename: {
                    **{
                        "rawFileName": raw_filename,
                        "filetype": ext_mapping[extension],
                        "filename": data.filename,
                        },
                    **metadata_contents.get("file_metadata", {}),
                }
            }
        })
        metadata_patch_response = patch_annotation(
            indicator_id=indicator_id,
            payload=annotation_payload
        )
        completed.append("file_metadata")

        job_string = "dataset_register_processors.finish_dataset_registration"
        job_id = f"{indicator_id}_{job_string}"

        context = get_context(indicator_id)

        job_data = q.enqueue_call(
            func=job_string,
            args=[context],
            kwargs={"filename": raw_filename},
            job_id=job_id,
        )

        completed.append("register_processor")

        job_status = job(uuid=indicator_id, job_string=job_string)
        completed.append("job_status")

    except KeyError as e:
        logger.error(f"Error parsing dictionary file. Attribute: {e}")
        errors = {"errors": [f"Dictionary file contains an invalid value: {e}"]}
        return augment_errors(errors, indicator_id, api_url)
    except HTTPException as e:
        logger.error(f"HTTPException (possible validation error) while running dataset register endpoint:\n{e}")
        errors = {"errors": [f"Dictionary file problem: {e.__cause__}"]}
        return augment_errors(errors, indicator_id, api_url)
    except Exception as ex_all:
        logger.error(f"Unexpected Exception while running dataset register endpoint:\n{ex_all}")
        errors = {"errors": [f"{ex_all}"]}
        return augment_errors(errors, indicator_id, api_url)
    finally:
        logger.info(f"Completed the following dataset register tasks: {completed}")

    del job_status["id"]

    return {
        **indicator_body,
        "job": {
            **job_status,
            "details": "Your dataset is being processed. Initial metadata has been uploaded, but additional processing time may be required for larger files. Use the indicator endpoint using the new dataset ID provided, and verify that the <published> property is set to <true>, which will indicate processing completion."
        }
    }


@router.get("/indicators/register/template")
def download_metadata_template_file():
    file_name = "dataset_register_metadata_template.json"
    headers = {
        "Content-Disposition": f"attachment; filename=dataset_metadata_template.json"
    }
    return FileResponse(file_name, headers=headers)


def bytes_to_csv(file):
    csv_reader = csv.DictReader(codecs.iterdecode(file, "utf-8"))
    return list(csv_reader)


@router.post("/indicators/{indicator_id}/annotations/file")
async def upload_data_dictionary_file(indicator_id: str, file: UploadFile = File(...)):
    """
    Accepts a CSV dictionary file describing a dataset in order to register it. Similar to using the API directly with JSON, or using the Dataset Registration flow on Dojo user interface to annotate a dataset.
    """

    if file.filename.endswith(".xlsx"):
        f = await file.read()
        xlsx = io.BytesIO(f)
        csv_dictionary_list = xls_to_annotations(xlsx)
    else:
        csv_dictionary_list = bytes_to_csv(file.file)

    try:
        formatted = format_annotations(csv_dictionary_list)
    except ValidationError as e:
        full_data = json.loads(e.json())

        try:
            full_data[0]["input_value"] = e.values
        # We are attaching e.values on our `format_to_schema` fn but handle
        # `values` not present in case we caught another ValidationError
        except AttributeError:
            pass

        full_data[0]["message"] = str(e)
        raise HTTPException(status_code=422, detail=full_data) from e

    annotation_payload = MetadataSchema.MetaModel(annotations=formatted)

    return patch_annotation(payload=annotation_payload, indicator_id=indicator_id)


@router.get("/indicators/annotations/file-template")
def download_data_dictionary_template_file(indicator_id=None, filetype="xlsx"):
    if filetype == "csv":
        file_name = "dataset_annotate_template.template_csv"
        headers = {
            "Content-Disposition": f"attachment; filename={file_name.replace('template_', '')}"
        }
        return FileResponse(file_name, headers=headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dojo Annotation Template"

    bold_font = Font(bold=True)

    columns = [
        {
            "name": "field_name",
            "help_text": """The name of the field as it exists in the source document.
This is usually the topmost cell in a column of a spreadsheet.""",
            "validation": None,
        },
        {
            "name": "group",
            "help_text": """Setting this allows you to group related fields such as multi-part dates, or multipart geos.
If the field is not part of a group, then this should be left blank.""",
            "validation": None,
        },
        {
            "name": "display_name",
            "help_text": """A name to display instead of field_name.""",
            "validation": None,
        },
        {
            "name": "description",
            "help_text": """Describes the field to provide better context during usage.""",
            "validation": None,
        },
        {
            "name": "data_type",
            "help_text": """Describes the type of data contained in the column.
Some options include:
integer - a numerical value that does not contain a decimal point/place
string - a word or text value that is not meant to be processed nor interpreted as a any of the other values available to describe the data.
float: a numerical value with no decimal place
binary: binary data in the dataset
boolean: data that represents yes/no values (true or false, enabled or disabled, etc)
""",
            "validation": DataValidation(
                type="list",
                formula1='"integer,string,float,binary,boolean,latitude,longitude,coordinates,country,iso2,iso3,state,territory,county,district,municipality,town,month,day,year,epoch,date"',
                allow_blank=False,
            ),
        },
        {
            "name": "units",
            "help_text": """
            Only include, and required, when data_type is not a value that directly correlates to time or location.
            """,
            "validation": None,
        },
        {
            "name": "units_description",
            "help_text": """An optional description for the feature unit, if the unit is required.""",
            "validation": None,
        },
        {
            "name": "primary",
            "help_text": """When data_type can be correlated to time or location (eg coordinates, day, year, etc), set only one location and one time as primary. Fields that are grouped together using the group column should all have the same value for this column.""",
            "validation": None,
            "validation": DataValidation(
                type="list",
                formula1='"Y,N"',
                allow_blank=True,
            ),
        },
        {
            "name": "date_format",
            "help_text": """Required when prividing a time data_type of month, day, year, or date. Python-compatible formatters only. See Dojo UI dataset registration for further help and examples.""",
            "validation": None,
        },
        {
            "name": "gadm_level",
            "help_text": """Level to map coordinates to. In other words, if a coordinate accuracy is at the country, state, territory (etc), level.""",
            "validation": None,
            "validation": DataValidation(
                type="list",
                formula1='"country,admin0,admin1,admin2,admin3"',
                allow_blank=True,
            ),
        },
        {
            "name": "resolve_to_gadm",
            "help_text": """If a location data_type field should be auto-resolved by the system to gadm.""",
            "validation": DataValidation(
                type="list",
                formula1='"Y,N"',
                allow_blank=True,
            ),
        },
        {
            "name": "coord_format",
            "help_text": """
            Only include, and required, when data_type is of coordinates. Some systems provide coordinates in latlon format, while others in lonlat. Please specify here, if applicable.
            """,
            "validation": None,
            "validation": DataValidation(
                type="list",
                formula1='"lonlat,latlon"',
                allow_blank=False,
            ),
        },
        {
            "name": "qualifies",
            "help_text": """""",
            "validation": None,
        },
        {
            "name": "qualifier_role",
            "help_text": """""",
            "validation": DataValidation(
                type="list",
                formula1='"breakdown,weight,minimum,maximum,coefficient"',
                allow_blank=False,
            ),
        },
    ]

    for index, col in enumerate(columns, start=1):
        col_name = col["name"]
        col_letter = openpyxl.utils.cell.get_column_letter(index)
        cell = ws.cell(row=1, column=index, value=col_name)
        cell.font = bold_font
        help_text = col.get("help_text", "").strip()
        help_prompt = DataValidation(
            type="custom",
            formula1="1",
            promptTitle=col_name,
            prompt=help_text,
            allow_blank=False,
            showInputMessage=True,
        )
        help_prompt.add(f"{col_letter}1")
        ws.add_data_validation(help_prompt)

        validation = col.get("validation", None)
        if validation:
            logger.warn(f"{col_name}: {validation}")
            validation.add(f"{col_letter}2:{col_letter}1048576")
            ws.add_data_validation(validation)

    if indicator_id:
        annotations = get_annotations(indicator_id)
        for index, row in enumerate(
            annotations.get("metadata", {}).get("column_statistics", {}).keys(), start=2
        ):
            ws.cell(row=index, column=1, value=row)

    # Freeze the top and left-most row/column
    ws.freeze_panes = "B2"

    file_name = "dataset_annotation_template.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename={file_name}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    from tempfile import NamedTemporaryFile

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        content = tmp.read()
    return Response(content=content, headers=headers)
